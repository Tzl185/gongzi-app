from openpyxl import load_workbook
import os
import pandas as pd

def read_excel_with_merged_header(filepath, header_row=4):
    """
    使用 openpyxl 解析第 header_row 行的合并表头，然后用 pandas 加载数据。
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = []
    for col in ws.iter_cols(min_row=header_row, max_row=header_row):
        cell = col[0]
        value = str(cell.value).strip() if cell.value else f"列{cell.column_letter}"
        headers.append(value)

    # 用pandas读取数据内容，跳过表头之前的行
    df = pd.read_excel(filepath, header=None, skiprows=header_row)
    df.columns = headers
    return df

def process_file_a(folder_path, output_file="文件A_汇总结果.xlsx"):
    all_data = []
    all_values = {}

    for filename in os.listdir(folder_path):
        if filename.endswith(('.xls', '.xlsx')) and not filename.startswith('~$'):
            filepath = os.path.join(folder_path, filename)
            try:
                print(f"\n🔍 正在处理: {filename}")
                df = read_excel_with_merged_header(filepath, header_row=4)

                budget_unit_col = df.columns[1]  # 第2列
                wage_cols = df.columns[16:30]   # 第17到30列

                df_filtered = df[[budget_unit_col] + list(wage_cols)]
                df_filtered[wage_cols] = df_filtered[wage_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                df_grouped = df_filtered.groupby(budget_unit_col).sum()

                for budget_unit, row in df_grouped.iterrows():
                    for wage_type in wage_cols:
                        value = row[wage_type]
                        wage_type_original = wage_type.strip()
                        if "绩效工资" in wage_type:
                            wage_type = wage_type.replace("绩效工资", "基础性绩效")
                        if "行政医疗" in wage_type:
                            wage_type = wage_type.replace("行政医疗", "职工基本医疗（行政）")
                        elif "事业医疗" in wage_type:
                            wage_type = wage_type.replace("事业医疗", "基本医疗（事业）")
                        elif "医疗保险" in wage_type:
                            wage_type = wage_type.replace("医疗保险", "基本医疗")

                        key = (str(budget_unit).strip(), str(wage_type).strip())
                        all_values[key] = value

                if not df_grouped.empty:
                    all_data.append(df_grouped)

            except Exception as e:
                print(f"❌ 文件 {filename} 处理失败: {e}")

    if all_data:
        df_all = pd.concat(all_data)
        df_final = df_all.groupby(df_all.index).sum()

        output_path = os.path.join(folder_path, output_file)
        df_final.to_excel(output_path)
        print(f"\n✅ 汇总结果已保存到: {output_path}")
        return output_path, all_values
    else:
        print("⚠️ 没有找到有效数据")
        return None, None

def update_file_b(file_a_path, file_b_path):
    """
    用文件A中的所有数值更新文件B的J列，保留原有格式
    """
    try:
        # 1. 从文件A中读取数据
        df_a = pd.read_excel(file_a_path, index_col=0)
        wage_cols = df_a.columns
        
        # 提取所有数值
        all_values = {}
        for budget_unit, row in df_a.iterrows():
            for wage_type in wage_cols:
                value = row[wage_type]
                if "绩效工资" in wage_type:
                    wage_type = wage_type.replace("绩效工资", "基础性绩效")
                if "行政医疗" in wage_type:
                    wage_type = wage_type.replace("行政医疗", "职工基本医疗（行政）")
                elif "事业医疗" in wage_type:  # 使用elif避免重复替换
                    wage_type = wage_type.replace("事业医疗", "基本医疗（事业）")
                elif "医疗保险" in wage_type:  # 可能还有其他表述
                    wage_type = wage_type.replace("医疗保险", "基本医疗")
                key = (str(budget_unit).strip(), str(wage_type).strip())
                all_values[key] = value

        # 2. 使用openpyxl直接操作Excel文件
        wb = load_workbook(file_b_path)
        sheet = wb.active
        
        # J列的索引（从1开始计数）
        j_col_index = 10
        
        # 3. 更新J列数据
        match_count = 0
        for row_idx in range(2, sheet.max_row + 1):  # 从第2行开始
            unit_cell = sheet.cell(row=row_idx, column=1)
            unit_info = str(unit_cell.value).strip() if unit_cell.value else ""
            
            budget_cell = sheet.cell(row=row_idx, column=2)
            budget_project = str(budget_cell.value).strip() if budget_cell.value else ""
            
            # 清理单位信息
            unit_info_cleaned = unit_info.replace("-", "").replace(" ", "")
            
            # 查找匹配
            matched = False
            for (budget_unit, wage_type), value in all_values.items():
                budget_unit_cleaned = budget_unit.replace("-", "").replace(" ", "")
                
                # 匹配条件
                unit_match = (budget_unit_cleaned in unit_info_cleaned) or (unit_info_cleaned in budget_unit_cleaned)
                wage_match = wage_type in budget_project
                
                if unit_match and wage_match:
                    # 更新单元格值，保留原有样式
                    sheet.cell(row=row_idx, column=j_col_index).value = value
                    match_count += 1
                    matched = True
                    print(f"匹配成功: 行{row_idx} 单位:'{unit_info}'⊇'{budget_unit}', 项目:'{budget_project}'⊇'{wage_type}', 值:{value}")
                    break
            
            if not matched and row_idx < 7:  # 打印前5行未匹配情况
                print(f"未匹配: 行{row_idx} 单位:'{unit_info}', 项目:'{budget_project}'")
        
        # 4. 保存更新后的文件B
        output_path = os.path.join(os.path.dirname(file_b_path), "updated_" + os.path.basename(file_b_path))
        wb.save(output_path)
        print(f"\n总共完成 {match_count} 处匹配")
        print(f"已保存更新后的文件B到: {output_path}")
        return output_path
    
    except Exception as e:
        print(f"\n更新文件B出错: {e}")
        return None

def main():
    # 文件夹路径（包含要处理的Excel表）
    folder_a = r"D:\副本\Desktop\合并表"
    # 文件B路径（模板文件）
    file_b_path = r"D:\副本\Desktop\项目细化导入模板-正数.xlsx"
    
    # 第一步：生成文件A
    print("第一步：处理文件夹生成文件A...")
    file_a_path, all_values = process_file_a(folder_a)
    
    if file_a_path and all_values:
        print("\n文件A中的数值示例（前5个）:")
        for i, (k, v) in enumerate(all_values.items()):
            if i < 5:
                print(f"{k}: {v}")
        
        # 第二步：用文件A更新文件B
        print("\n第二步：用文件A更新文件B...")
        update_file_b(file_a_path, file_b_path)
    else:
        print("未能生成有效的文件A，请检查输入文件")

if __name__ == "__main__":
    main()
